import openpyxl
from collections import Counter

wb = openpyxl.load_workbook("双色球近50期开奖结果.xlsx")
sh = wb.worksheets[0]
RAW_DATA = []
for r in range(5, sh.max_row + 1):
    summary = sh.cell(row=r, column=7).value
    if not summary: continue
    parts = summary.replace('+', ' ').split()
    try:
        nums = [int(p) for p in parts]
        RAW_DATA.append((str(sh.cell(row=r, column=2).value), nums[:6], nums[6]))
    except: pass

ALL_RED = list(range(1, 34))
ALL_BLUE = list(range(1, 17))
Z1, Z2, Z3 = [r for r in ALL_RED if r <= 11], [r for r in ALL_RED if 12 <= r <= 22], [r for r in ALL_RED if r >= 23]

def top_n_in_zone(scores, zone, n):
    return sorted([k for k in zone], key=lambda k: scores.get(k, 0), reverse=True)[:n]

def m12(red_hist, blue_hist, n_train, alpha=0.92):
    r_cnt = Counter()
    for i in range(n_train):
        w = alpha ** (n_train - 1 - i)
        for r in red_hist[i]: r_cnt[r] += w
    rmax = max(r_cnt.values()) if r_cnt else 1
    rs = {n: r_cnt.get(n, 0) / rmax * 100 for n in ALL_RED}
    bs = {n: 10 for n in ALL_BLUE}
    lb = blue_hist[n_train - 1]
    if lb > 1: bs[lb - 1] = 100
    if lb < 16: bs[lb + 1] = 100
    return rs, bs

all_reds = [r for _, r, _ in RAW_DATA]
all_blues = [b for _, _, b in RAW_DATA]

total_rh, total_bh = 0, 0
lines = []

print(f"{'期号':>8} | {'类型':<4} | {'一区(01-11)':<22} | {'二区(12-22)':<22} | {'三区(23-33)':<22} | {'蓝球':<10} | {'红中':>3} | {'蓝中':>3}")
print("-" * 120)

for test_idx in range(10, len(RAW_DATA)):
    period, actual_reds, actual_blue = RAW_DATA[test_idx]
    z1a = sorted([r for r in actual_reds if r <= 11])
    z2a = sorted([r for r in actual_reds if 12 <= r <= 22])
    z3a = sorted([r for r in actual_reds if r >= 23])
    
    r_scores, b_scores = m12(all_reds, all_blues, test_idx)
    z1p = sorted(top_n_in_zone(r_scores, Z1, 4))
    z2p = sorted(top_n_in_zone(r_scores, Z2, 4))
    z3p = sorted(top_n_in_zone(r_scores, Z3, 4))
    bp = sorted(sorted(b_scores, key=b_scores.get, reverse=True)[:4])
    
    rh = len(set(z1p) & set(z1a)) + len(set(z2p) & set(z2a)) + len(set(z3p) & set(z3a))
    bh = 1 if actual_blue in bp else 0
    total_rh += rh
    total_bh += bh
    
    z1ps = ' '.join(f'{n:02d}' for n in z1p) if z1p else '—'
    z2ps = ' '.join(f'{n:02d}' for n in z2p) if z2p else '—'
    z3ps = ' '.join(f'{n:03d}' for n in z3p) if z3p else '—'
    bps  = ' '.join(f'{n:02d}' for n in bp)
    
    z1as = ' '.join(f'{n:02d}' for n in z1a) if z1a else '—'
    z2as = ' '.join(f'{n:02d}' for n in z2a) if z2a else '—'
    z3as = ' '.join(f'{n:02d}' for n in z3a) if z3a else '—'
    
    # 第1行：预测
    print(f"{period:>8} | {'预测':<4} | {z1ps:<22} | {z2ps:<22} | {z3ps:<22} | {bps:<10} |     |")
    # 第2行：实际
    print(f"{'':>8} | {'实际':<4} | {z1as:<22} | {z2as:<22} | {z3as:<22} | {actual_blue:02d}       | {rh:>3} | {'Y' if bh else '-'}")
    print("-" * 120)

n = len(RAW_DATA) - 10
print()
print(f"汇总 | 红球总命中 {total_rh}/240 = {total_rh/240*100:.1f}% | 蓝球总命中 {total_bh}/40 = {total_bh/40*100:.1f}% | 每期均中 {total_rh/n:.2f} 个")
