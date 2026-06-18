# -*- coding: utf-8 -*-
"""生成双色球预测结果Excel"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== 样式 =====
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
sub_header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
sub_header_font = Font(name="微软雅黑", size=11, bold=True, color="2F5496")
data_font = Font(name="微软雅黑", size=11)
title_font = Font(name="微软雅黑", size=16, bold=True, color="1F3864")
hot_font = Font(name="微软雅黑", size=12, bold=True, color="CC0000")
recommend_font = Font(name="微软雅黑", size=13, bold=True, color="FFFFFF")
recommend_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
zone1_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
zone2_fill = PatternFill(start_color="FFFDE0", end_color="FFFDE0", fill_type="solid")
zone3_fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
blue_fill = PatternFill(start_color="E0E8FF", end_color="E0E8FF", fill_type="solid")
gold_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
green_fill_score = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
thin_border = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ============================================================
# Sheet 1: 预测总览
# ============================================================
ws = wb.active
ws.title = "预测总览"

# 标题
ws.merge_cells('A1:H1')
ws['A1'] = '双色球下期号码预测报告'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

ws.merge_cells('A2:H2')
ws['A2'] = '基于近50期(2026017~2026066)数据 | 多模型综合预测 | 分析时间: 2026-06-15'
ws['A2'].font = Font(name="微软雅黑", size=9, color="888888")
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

# 上期开奖
ws.merge_cells('A4:H4')
ws['A4'] = '【上期开奖】第2026066期: 红球 05 11 21 23 24 29 + 蓝球 16'
ws['A4'].font = Font(name="微软雅黑", size=11, bold=True, color="333333")
ws['A4'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ws['A4'].alignment = center_align
ws.row_dimensions[4].height = 28

# 红球推荐
row = 6
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = '🔴 红球备选推荐（每个区间4个号码）'
ws[f'A{row}'].font = Font(name="微软雅黑", size=14, bold=True, color="CC0000")
ws.row_dimensions[row].height = 32

row = 8
headers = ['区间', '首选', '次选', '备选1', '备选2', '综合得分趋势', '选号建议', '推荐数量']
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 一区数据
zone_data = [
    ('一区 (01-11)', ['02', '01', '03', '10'], [71.5, 57.3, 57.0, 50.5],
     '02为近50期最热号之一，\n01近期频繁出现，\n03和10有连号/邻号支撑', '选2~3个',
     PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")),
    ('二区 (12-22)', ['22', '18', '15', '14'], [61.6, 51.4, 51.2, 48.8],
     '22为二区最强热号，\n18连号模式频繁，\n14-15为邻居组合', '选1~2个',
     PatternFill(start_color="FFFDE0", end_color="FFFDE0", fill_type="solid")),
    ('三区 (23-33)', ['30', '29', '25', '24'], [77.9, 76.2, 66.9, 64.0],
     '30为最强热号(出现16次)，\n29/25/24均高热+邻号，\n29-30为经典连号组合', '选2~3个',
     PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")),
]

for i, (zone, nums, scores, reason, count, fill) in enumerate(zone_data):
    r = row + 1 + i
    ws.cell(row=r, column=1, value=zone).font = Font(name="微软雅黑", size=11, bold=True)
    ws.cell(row=r, column=1).fill = fill
    ws.cell(row=r, column=1).alignment = center_align
    ws.cell(row=r, column=1).border = thin_border
    
    for j, (num, score) in enumerate(zip(nums, scores)):
        cell = ws.cell(row=r, column=2+j, value=f"{num} ({score:.1f}分)")
        cell.font = Font(name="微软雅黑", size=12, bold=True, color="CC0000" if j==0 else "333333")
        cell.fill = fill if j == 0 else gold_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    ws.cell(row=r, column=6, value=reason).font = Font(name="微软雅黑", size=9)
    ws.cell(row=r, column=6).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.cell(row=r, column=6).border = thin_border
    
    ws.cell(row=r, column=7, value=count).font = Font(name="微软雅黑", size=11, bold=True)
    ws.cell(row=r, column=7).alignment = center_align
    ws.cell(row=r, column=7).border = thin_border
    
    # 推荐数量用红色框标注
    ws.cell(row=r, column=8, value=count).font = Font(name="微软雅黑", size=12, bold=True, color="CC0000")
    ws.cell(row=r, column=8).alignment = center_align
    ws.cell(row=r, column=8).border = thin_border
    ws.cell(row=r, column=8).fill = fill
    
    ws.row_dimensions[r].height = 55

# 蓝球推荐
row = 13
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = '🔵 蓝球备选推荐（4个号码）'
ws[f'A{row}'].font = Font(name="微软雅黑", size=14, bold=True, color="0044CC")
ws.row_dimensions[row].height = 32

row = 15
blue_headers = ['号码', '综合得分', '奇偶', '大小', '热度', '分析理由']
for i, h in enumerate(blue_headers, 1):
    cell = ws.cell(row=row, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

blue_data = [
    ('07', 59.4, '奇数', '小(1-8)', '🔥热号(5次)', '奇小组合+遗漏值低+马尔可夫转移高概率+邻号优势'),
    ('02', 51.5, '偶数', '小(1-8)', '🔥热号(8次)', '最高频蓝球+最近2期内刚出现+小号回补趋势'),
    ('01', 44.7, '奇数', '小(1-8)', '🔥热号(4次)', '奇偶交替规律+小号密集+遗漏不高'),
    ('11', 43.9, '奇数', '大(9-16)', '⚡温号(3次)', '奇数延续+大号反弹+跨度适中(距上期蓝16为5)'),
]

for i, (num, score, oe, bs, hot, reason) in enumerate(blue_data):
    r = row + 1 + i
    ws.cell(row=r, column=1, value=num).font = Font(name="微软雅黑", size=14, bold=True, color="0044CC")
    ws.cell(row=r, column=1).fill = blue_fill
    ws.cell(row=r, column=1).alignment = center_align
    ws.cell(row=r, column=1).border = thin_border
    
    ws.cell(row=r, column=2, value=f"{score:.1f}").font = Font(name="微软雅黑", size=12, bold=True)
    ws.cell(row=r, column=2).alignment = center_align
    ws.cell(row=r, column=2).border = thin_border
    if i == 0:
        ws.cell(row=r, column=2).fill = gold_fill
    
    for j, val in enumerate([oe, bs, hot, reason], 3):
        cell = ws.cell(row=r, column=j, value=val)
        cell.font = data_font
        cell.alignment = center_align if j < 6 else Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[r].height = 35

# 推荐组合
row = 21
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = '⭐ 推荐选号策略: 区间比 2:1:3 (一区2个 + 二区1个 + 三区3个)'
ws[f'A{row}'].font = Font(name="微软雅黑", size=13, bold=True, color="333333")
ws[f'A{row}'].fill = gold_fill
ws[f'A{row}'].alignment = center_align
ws.row_dimensions[row].height = 30

row = 22
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = '示例组合: 02 01 | 22 | 30 29 25  +  07'
ws[f'A{row}'].font = Font(name="微软雅黑", size=14, bold=True, color="CC0000")
ws[f'A{row}'].alignment = center_align
ws.row_dimensions[row].height = 32

# 免责
row = 24
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = '⚠️ 彩票开奖为完全随机事件，以上预测基于历史数据统计模型，仅供娱乐参考。请理性购彩！'
ws[f'A{row}'].font = Font(name="微软雅黑", size=9, color="999999")
ws[f'A{row}'].alignment = center_align

# 列宽
widths = {1: 16, 2: 14, 3: 14, 4: 14, 5: 14, 6: 32, 7: 14, 8: 12}
for k, v in widths.items():
    ws.column_dimensions[get_column_letter(k)].width = v

# ============================================================
# Sheet 2: 详细评分表
# ============================================================
ws2 = wb.create_sheet("详细评分")

ws2.merge_cells('A1:I1')
ws2['A1'] = '红球各号码多模型评分明细'
ws2['A1'].font = Font(name="微软雅黑", size=14, bold=True, color="333333")
ws2['A1'].alignment = center_align

headers2 = ['号码', '区间', '频率(20%)', '遗漏(15%)', '邻号(15%)', '马尔可夫(15%)', '贝叶斯(5%)', '连号(15%)', '跨度(15%)', '综合得分', '评级']
for i, h in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 红球评分数据
red_scores = [
    (1, '一区', 81.2, 91.7, 7.1, 50.5, 5.0, 42.9, 80.0, 57.3, '★★★☆'),
    (2, '一区', 87.5, 87.5, 40.0, 100.0, 2.5, 71.4, 60.0, 71.5, '★★★★'),
    (3, '一区', 75.0, 45.8, 4.8, 71.4, 2.5, 57.1, 100.0, 57.0, '★★★☆'),
    (4, '一区', 43.8, 83.3, 7.1, 41.5, 4.2, 42.9, 40.0, 41.2, '★★★'),
    (5, '一区', 18.8, 100.0, 0.0, 3.2, 1.7, 14.3, 20.0, 24.5, '★☆'),
    (6, '一区', 43.8, 33.3, 20.0, 17.1, 0.8, 0.0, 0.0, 19.4, '★☆'),
    (7, '一区', 50.0, 95.8, 20.0, 57.7, 4.2, 28.6, 30.0, 45.0, '★★★'),
    (8, '一区', 50.0, 95.8, 17.6, 42.4, 3.3, 28.6, 40.0, 43.8, '★★★'),
    (9, '一区', 56.2, 91.7, 11.8, 33.0, 4.2, 28.6, 30.0, 40.7, '★★★'),
    (10, '一区', 56.2, 75.0, 18.8, 54.3, 2.5, 42.9, 70.0, 50.5, '★★★☆'),
    (11, '一区', 50.0, 100.0, 30.8, 39.3, 2.5, 14.3, 10.0, 39.3, '★★★'),
    (12, '二区', 25.0, 0.0, 11.8, 20.3, 0.0, 14.3, 30.0, 16.5, '★'),
    (13, '二区', 62.5, 50.0, 23.1, 61.0, 1.7, 42.9, 30.0, 43.6, '★★★'),
    (14, '二区', 56.2, 83.3, 10.0, 48.7, 3.3, 57.1, 50.0, 48.8, '★★★☆'),
    (15, '二区', 62.5, 91.7, 17.6, 54.3, 4.2, 42.9, 50.0, 51.2, '★★★☆'),
    (16, '二区', 50.0, 95.8, 22.2, 36.8, 4.2, 28.6, 30.0, 42.2, '★★★'),
    (17, '二区', 50.0, 12.5, 26.3, 43.3, 0.0, 71.4, 80.0, 45.0, '★★★'),
    (18, '二区', 68.8, 91.7, 15.4, 35.8, 2.5, 57.1, 50.0, 51.4, '★★★☆'),
    (19, '二区', 31.2, 58.3, 0.0, 21.8, 0.8, 0.0, 20.0, 21.3, '★☆'),
    (20, '二区', 18.8, 50.0, 8.3, 17.4, 1.7, 0.0, 10.0, 16.7, '★'),
    (21, '二区', 50.0, 100.0, 6.2, 37.7, 3.3, 14.3, 40.0, 39.9, '★★★'),
    (22, '二区', 81.2, 75.0, 31.2, 77.9, 3.3, 57.1, 60.0, 61.6, '★★★★'),
    (23, '三区', 62.5, 100.0, 11.5, 55.1, 1.7, 71.4, 90.0, 61.8, '★★★★'),
    (24, '三区', 87.5, 100.0, 21.7, 69.2, 5.0, 57.1, 60.0, 64.0, '★★★★'),
    (25, '三区', 87.5, 87.5, 29.4, 74.0, 3.3, 57.1, 80.0, 66.9, '★★★★'),
    (26, '三区', 25.0, 70.8, 5.3, 25.7, 1.7, 28.6, 30.0, 29.1, '★★'),
    (27, '三区', 31.2, 75.0, 18.8, 16.2, 2.5, 14.3, 40.0, 31.0, '★★'),
    (28, '三区', 75.0, 87.5, 23.5, 63.0, 5.8, 57.1, 80.0, 62.0, '★★★★'),
    (29, '三区', 81.2, 100.0, 39.3, 68.2, 5.8, 100.0, 90.0, 76.2, '★★★★★'),
    (30, '三区', 100.0, 95.8, 28.6, 73.2, 7.5, 85.7, 100.0, 77.9, '★★★★★'),
    (31, '三区', 56.2, 87.5, 18.2, 37.9, 2.5, 28.6, 70.0, 47.7, '★★★'),
    (32, '三区', 37.5, 95.8, 21.1, 14.9, 1.7, 14.3, 20.0, 32.5, '★★'),
    (33, '三区', 62.5, 91.7, 16.7, 46.3, 4.2, 14.3, 50.0, 45.6, '★★★'),
]

zone_fills = {'一区': zone1_fill, '二区': zone2_fill, '三区': zone3_fill}
for i, row_data in enumerate(red_scores):
    r = 4 + i
    num, zone, *scores, composite, rating = row_data
    
    ws2.cell(row=r, column=1, value=f"{num:02d}").font = Font(name="微软雅黑", size=11, bold=True)
    ws2.cell(row=r, column=1).alignment = center_align
    ws2.cell(row=r, column=1).border = thin_border
    
    ws2.cell(row=r, column=2, value=zone).font = Font(name="微软雅黑", size=11, bold=True)
    ws2.cell(row=r, column=2).fill = zone_fills[zone]
    ws2.cell(row=r, column=2).alignment = center_align
    ws2.cell(row=r, column=2).border = thin_border
    
    for j, s in enumerate(scores):
        cell = ws2.cell(row=r, column=3+j, value=s)
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # 综合得分高亮
    comp_cell = ws2.cell(row=r, column=10, value=composite)
    comp_cell.font = Font(name="微软雅黑", size=12, bold=True, color="CC0000" if composite >= 60 else "333333")
    comp_cell.alignment = center_align
    comp_cell.border = thin_border
    if composite >= 60:
        comp_cell.fill = gold_fill
    
    ws2.cell(row=r, column=11, value=rating).font = Font(name="微软雅黑", size=11)
    ws2.cell(row=r, column=11).alignment = center_align
    ws2.cell(row=r, column=11).border = thin_border
    
    ws2.row_dimensions[r].height = 22

# 列宽
ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 8
for c in range(3, 10):
    ws2.column_dimensions[get_column_letter(c)].width = 12
ws2.column_dimensions['J'].width = 12
ws2.column_dimensions['K'].width = 10

ws2.freeze_panes = 'A4'

# ============================================================
# Sheet 3: 方法说明
# ============================================================
ws3 = wb.create_sheet("分析方法说明")

ws3.merge_cells('A1:C1')
ws3['A1'] = '预测方法说明'
ws3['A1'].font = Font(name="微软雅黑", size=14, bold=True, color="333333")
ws3['A1'].alignment = center_align

methods = [
    ('方法名称', '权重', '说明'),
    ('1. 频率分析 (Hot/Cold)', '红20%/蓝25%', '统计50期内每个号码出现次数，热号得分高'),
    ('2. 遗漏分析 (Missing)', '红15%/蓝20%', '距离上次出现期数越短得分越高，捕捉回补规律'),
    ('3. 邻号分析 (Adjacent)', '红15%/蓝15%', '上期号码±1的邻号历史出现概率（用户规律4）'),
    ('4. 马尔可夫链 (Markov)', '红15%/蓝15%', '基于上期号码集合，计算下期各号码的转移概率'),
    ('5. 贝叶斯条件概率', '红5%/蓝10%', '基于区间分布模式，计算条件概率'),
    ('6. 连号分析 (Consecutive)', '红15%/-', '统计连号对出现的频率，提高连号候选得分（用户规律5）'),
    ('7. 跨度分析 (Gap)', '红15%/蓝15%', '分析相邻号码间距，小区间号码得分高（用户规律6）'),
]

for i, (m, w, d) in enumerate(methods):
    r = 3 + i
    for j, val in enumerate([m, w, d], 1):
        cell = ws3.cell(row=r, column=j, value=val)
        if i == 0:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = data_font
        cell.alignment = Alignment(horizontal='left' if j==3 else 'center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws3.row_dimensions[r].height = 28

# 用户规律对应
ws3.merge_cells('A12:C12')
ws3['A12'] = '用户总结的规律与模型对应'
ws3['A12'].font = Font(name="微软雅黑", size=12, bold=True, color="333333")

user_rules = [
    ('用户规律', '对应模型', '处理方式'),
    ('规律1: 每区1~3个号码', '区间比模式分析', '分析近10期趋势，推荐2:1:3区间比'),
    ('规律2: 热号横热', '频率分析', '频率权重最高(20%)，热号得分高'),
    ('规律3: 隔2-3期热号频率高', '遗漏分析', '遗漏值越小得分越高，捕捉周期性回补'),
    ('规律4: 上期周围号码概率高', '邻号分析', '上期号码±1邻号加权评分'),
    ('规律5: 连号情况', '连号分析', '历史连号对频率统计，提升连号候选'),
    ('规律6: 相邻间距不大', '跨度分析', '小区间(≤2)涉及的号码加分'),
]

for i, (rule, model, method) in enumerate(user_rules):
    r = 14 + i
    for j, val in enumerate([rule, model, method], 1):
        cell = ws3.cell(row=r, column=j, value=val)
        if i == 0:
            cell.font = Font(name="微软雅黑", size=11, bold=True)
            cell.fill = sub_header_fill
        else:
            cell.font = data_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws3.row_dimensions[r].height = 30

ws3.column_dimensions['A'].width = 32
ws3.column_dimensions['B'].width = 22
ws3.column_dimensions['C'].width = 45

# 保存
output = r"C:\Users\27218\WorkBuddy\2026-06-12-16-55-40\双色球预测报告.xlsx"
wb.save(output)
print(f"Excel保存到: {output}")
