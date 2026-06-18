"""
提取S1最佳方法(指数衰减+蓝球邻号)逐期预测vs实际对照表
"""
import openpyxl
from collections import Counter

wb = openpyxl.load_workbook("双色球近50期开奖结果.xlsx")
sh = wb.worksheets[0]
RAW_DATA = []
for r in range(5, sh.max_row + 1):
    summary = sh.cell(row=r, column=7).value
    if not summary:
        continue
    parts = summary.replace('+', ' ').split()
    try:
        nums = [int(p) for p in parts]
        period = str(sh.cell(row=r, column=2).value)
        RAW_DATA.append((period, nums[:6], nums[6]))
    except:
        pass

ALL_RED = list(range(1, 34))
ALL_BLUE = list(range(1, 17))
Z1 = [r for r in ALL_RED if r <= 11]
Z2 = [r for r in ALL_RED if 12 <= r <= 22]
Z3 = [r for r in ALL_RED if r >= 23]

def top_n_in_zone(scores, zone, n):
    zs = {k: scores.get(k, 0) for k in zone}
    return sorted(zs, key=zs.get, reverse=True)[:n]

def m12_exponential_decay(red_hist, blue_hist, n_train, alpha=0.92):
    r_cnt = Counter()
    for i in range(n_train):
        w = alpha ** (n_train - 1 - i)
        for r in red_hist[i]:
            r_cnt[r] += w
    rmax = max(r_cnt.values()) if r_cnt else 1
    rs = {n: r_cnt.get(n, 0) / rmax * 100 for n in ALL_RED}
    bs = {n: 10 for n in ALL_BLUE}
    lb = blue_hist[n_train - 1]
    if lb > 1: bs[lb - 1] = 100
    if lb < 16: bs[lb + 1] = 100
    return rs, bs

all_reds = [r for _, r, _ in RAW_DATA]
all_blues = [b for _, _, b in RAW_DATA]

print(f"{'期号':>8} | {'一区预测':<20} | {'二区预测':<20} | {'三区预测':<20} | {'蓝球预测'} | {'一区实际':<16} | {'二区实际':<16} | {'三区实际':<16} | {'蓝球实际'} | {'红中':>3} | {'蓝中'}")
print("-" * 160)

total_rh = 0
total_bh = 0

for test_idx in range(10, len(RAW_DATA)):
    period, actual_reds, actual_blue = RAW_DATA[test_idx]
    actual = set(actual_reds)
    z1a = [r for r in actual_reds if r <= 11]
    z2a = [r for r in actual_reds if 12 <= r <= 22]
    z3a = [r for r in actual_reds if r >= 23]
    
    r_scores, b_scores = m12_exponential_decay(all_reds, all_blues, test_idx)
    
    z1p = top_n_in_zone(r_scores, Z1, 4)
    z2p = top_n_in_zone(r_scores, Z2, 4)
    z3p = top_n_in_zone(r_scores, Z3, 4)
    bp = sorted(b_scores, key=b_scores.get, reverse=True)[:4]
    
    h1 = len(set(z1p) & set(z1a))
    h2 = len(set(z2p) & set(z2a))
    h3 = len(set(z3p) & set(z3a))
    rh = h1 + h2 + h3
    bh = 1 if actual_blue in bp else 0
    total_rh += rh
    total_bh += bh
    
    z1p_s = ' '.join(f'{n:02d}' for n in sorted(z1p))
    z2p_s = ' '.join(f'{n:02d}' for n in sorted(z2p))
    z3p_s = ' '.join(f'{n:02d}' for n in sorted(z3p))
    bp_s   = ' '.join(f'{n:02d}' for n in sorted(bp))
    
    z1a_s = ' '.join(f'{n:02d}' for n in sorted(z1a))
    z2a_s = ' '.join(f'{n:02d}' for n in sorted(z2a))
    z3a_s = ' '.join(f'{n:02d}' for n in sorted(z3a))
    
    print(f"{period:>8} | {z1p_s:<20} | {z2p_s:<20} | {z3p_s:<20} | {bp_s:<10} | {z1a_s:<16} | {z2a_s:<16} | {z3a_s:<16} | {actual_blue:02d}       | {rh:>3} | {'Y' if bh else '-'}")

n = len(RAW_DATA) - 10
print()
print(f"40期汇总：红球总命中 {total_rh}/240 = {total_rh/240*100:.1f}% | 蓝球总命中 {total_bh}/40 = {total_bh/40*100:.1f}%")
print(f"每期均中红球 {total_rh/n:.2f} 个")
