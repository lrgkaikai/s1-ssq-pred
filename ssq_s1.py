# -*- coding: utf-8 -*-
"""
S1 双色球预测系统 — 14方法独立回测（含定期洗牌分区选方法）
从零重写，无融合，无动态权重，纯独立方法对比
"""
import math
from collections import Counter

# ============================================================
RAW_DATA = [
    ("2026017", [1, 3, 5, 18, 29, 32], 4),
    ("2026018", [11, 15, 17, 22, 25, 30], 7),
    ("2026019", [7, 8, 16, 17, 18, 30], 1),
    ("2026020", [1, 13, 14, 21, 24, 30], 2),
    ("2026021", [3, 13, 25, 26, 30, 31], 4),
    ("2026022", [15, 18, 23, 25, 28, 32], 11),
    ("2026023", [1, 3, 8, 10, 23, 29], 6),
    ("2026024", [1, 2, 13, 21, 23, 29], 14),
    ("2026025", [2, 3, 15, 20, 23, 24], 10),
    ("2026026", [2, 9, 16, 22, 25, 29], 3),
    ("2026027", [2, 13, 17, 18, 25, 26], 13),
    ("2026028", [2, 6, 9, 17, 25, 28], 15),
    ("2026029", [6, 19, 22, 23, 28, 31], 5),
    ("2026030", [10, 11, 14, 19, 22, 24], 4),
    ("2026031", [3, 10, 12, 13, 18, 33], 8),
    ("2026032", [1, 3, 11, 18, 31, 33], 2),
    ("2026033", [3, 6, 13, 21, 28, 29], 6),
    ("2026034", [1, 3, 7, 13, 22, 23], 7),
    ("2026035", [2, 6, 12, 24, 25, 32], 2),
    ("2026036", [6, 10, 12, 15, 22, 28], 8),
    ("2026037", [11, 22, 27, 29, 31, 33], 12),
    ("2026038", [1, 2, 13, 23, 25, 27], 5),
    ("2026039", [8, 17, 18, 21, 25, 30], 5),
    ("2026040", [3, 4, 14, 22, 23, 33], 4),
    ("2026041", [2, 8, 10, 17, 19, 24], 13),
    ("2026042", [2, 7, 12, 19, 24, 31], 10),
    ("2026043", [6, 9, 14, 16, 25, 32], 16),
    ("2026044", [2, 14, 17, 18, 22, 30], 1),
    ("2026045", [4, 11, 15, 17, 24, 30], 15),
    ("2026046", [2, 9, 10, 24, 31, 33], 16),
    ("2026047", [7, 16, 21, 24, 27, 30], 7),
    ("2026048", [9, 15, 18, 24, 28, 33], 1),
    ("2026049", [3, 4, 14, 15, 18, 20], 2),
    ("2026050", [6, 9, 25, 27, 28, 30], 3),
    ("2026051", [9, 14, 15, 16, 29, 30], 10),
    ("2026052", [1, 3, 11, 22, 26, 31], 11),
    ("2026053", [1, 2, 3, 8, 13, 14], 2),
    ("2026054", [13, 20, 25, 29, 30, 33], 2),
    ("2026055", [4, 11, 24, 25, 32, 33], 13),
    ("2026056", [10, 19, 21, 22, 31, 33], 5),
    ("2026057", [1, 10, 22, 24, 28, 30], 7),
    ("2026058", [1, 4, 7, 21, 29, 30], 1),
    ("2026059", [8, 16, 26, 28, 29, 30], 15),
    ("2026060", [7, 9, 10, 16, 22, 27], 11),
    ("2026061", [1, 4, 5, 15, 23, 28], 7),
    ("2026062", [2, 4, 7, 14, 28, 29], 9),
    ("2026063", [2, 8, 25, 28, 30, 31], 2),
    ("2026064", [1, 9, 15, 18, 29, 33], 15),
    ("2026065", [7, 8, 16, 24, 30, 32], 2),
    ("2026066", [5, 11, 21, 23, 24, 29], 16),
    ("2026067", [4, 19, 27, 29, 30, 32], 13),
    ("2026068", [3, 5, 16, 18, 29, 32], 4),
]
N = len(RAW_DATA)  # 52

ALL_RED  = list(range(1, 34))
ALL_BLUE = list(range(1, 17))
Z1 = list(range(1, 12))
Z2 = list(range(12, 23))
Z3 = list(range(23, 34))
ZONES = {"Z1": Z1, "Z2": Z2, "Z3": Z3}

# ============================================================
# 工具函数
# ============================================================
def top_n_in_zone(scores, zone, n=4):
    """从scores中选出zone内分数最高的n个号码（按分数降序）"""
    zs = [(num, scores.get(num, 0)) for num in zone]
    zs.sort(key=lambda x: -x[1])
    return [num for num, _ in zs[:n]]

def top_n_blue(scores, n=4):
    if not scores or len(set(scores.values())) <= 1:  # 全同分=无效
        return []
    bs = [(num, scores.get(num, 0)) for num in ALL_BLUE]
    bs.sort(key=lambda x: -x[1])
    return [num for num, _ in bs[:n]]

def norm01(d):
    """归一化到0-100"""
    mx = max(d.values()) if d else 1
    if mx == 0:
        return {k: 0 for k in d}
    return {k: v / mx * 100 for k, v in d.items()}

def zonal_norm(scores):
    """按区独立归一化"""
    result = {}
    for zn, zl in ZONES.items():
        zs = {n: scores.get(n, 0) for n in zl}
        mx = max(zs.values()) if zs else 1
        if mx > 0:
            for n in zl:
                result[n] = zs[n] / mx * 100
        else:
            for n in zl:
                result[n] = 0
    return result

# ============================================================
# 9个V1-V3基础方法 + 4个V5方法
# 每个方法签名为: (red_hist, blue_hist, train_size) → (red_scores, blue_scores)
# ============================================================

# ── 1. 频率分析 (冷热号) ──
def m01_frequency(red_hist, blue_hist, n_train):
    rcnt = Counter()
    for rs in red_hist[:n_train]:
        rcnt.update(rs)
    bcnt = Counter(blue_hist[:n_train])
    rmax = max(rcnt.values()) if rcnt else 1
    bmax = max(bcnt.values()) if bcnt else 1
    rs = {n: rcnt.get(n, 0) / rmax * 100 for n in ALL_RED}
    bs = {n: bcnt.get(n, 0) / bmax * 100 for n in ALL_BLUE}
    return rs, bs

# ── 2. 遗漏分析 ──
def m02_missing(red_hist, blue_hist, n_train):
    rlast = {}
    blast = {}
    for i in range(n_train):
        for r in red_hist[i]:
            rlast[r] = i
        blast[blue_hist[i]] = i
    rs = {}
    for n in ALL_RED:
        miss = n_train - 1 - rlast.get(n, -1)
        rs[n] = 100 / (miss + 1)
    bs = {}
    for n in ALL_BLUE:
        miss = n_train - 1 - blast.get(n, -1)
        bs[n] = 100 / (miss + 1)
    return zonal_norm(rs), norm01(bs)

# ── 3. 邻号分析 ──
def m03_adjacent(red_hist, blue_hist, n_train):
    if n_train < 2:
        return {n: 50 for n in ALL_RED}, {}
    # 红球邻号统计
    adj_hit = Counter()
    adj_tot = Counter()
    for i in range(1, n_train):
        for p in red_hist[i-1]:
            nb = set()
            if p > 1: nb.add(p-1)
            if p < 33: nb.add(p+1)
            for a in nb:
                adj_tot[a] += 1
                if a in red_hist[i]:
                    adj_hit[a] += 1
    last = red_hist[n_train-1]
    last_adj = set()
    for n in last:
        if n > 1: last_adj.add(n-1)
        if n < 33: last_adj.add(n+1)
    rs = {}
    for n in ALL_RED:
        base = adj_hit.get(n, 0) / max(adj_tot.get(n, 1), 1) * 100
        if n in last_adj:
            base += 30  # 降低加分，避免过度倾斜
        rs[n] = base
    # 蓝球邻号 — 纯邻号优先策略
    lb = blue_hist[n_train-1]
    la = set()
    if lb > 1: la.add(lb-1)
    if lb < 16: la.add(lb+1)
    bs = {}
    for n in ALL_BLUE:
        if n in la:
            bs[n] = 100  # 邻号绝对优先
        else:
            bs[n] = 10   # 非邻号低保
    return zonal_norm(rs), norm01(bs)

# ── 4. 马尔可夫链 ──
def m04_markov(red_hist, blue_hist, n_train):
    if n_train < 2:
        return {n: 50 for n in ALL_RED}, {}
    # 红球转移矩阵
    rtrans = {n: Counter() for n in ALL_RED}
    rtotal = {n: 0 for n in ALL_RED}
    for i in range(1, n_train):
        for p in red_hist[i-1]:
            rtotal[p] += len(red_hist[i])
            for c in red_hist[i]:
                rtrans[p][c] += 1
    rs = Counter()
    for seed in red_hist[n_train-1]:
        t = max(rtotal.get(seed, 0), 1)
        for n in ALL_RED:
            rs[n] += rtrans[seed].get(n, 0) / t * 100
    # 蓝球转移
    btrans = {n: Counter() for n in ALL_BLUE}
    btotal = {n: 0 for n in ALL_BLUE}
    for i in range(1, n_train):
        pb = blue_hist[i-1]
        cb = blue_hist[i]
        btotal[pb] += 1
        btrans[pb][cb] += 1
    bs = {}
    lb = blue_hist[n_train-1]
    t = max(btotal.get(lb, 0), 1)
    for n in ALL_BLUE:
        bs[n] = btrans[lb].get(n, 0) / t * 100
    return zonal_norm(dict(rs)), norm01(bs)

# ── 5. 贝叶斯区间比 ──
def m05_bayesian(red_hist, blue_hist, n_train):
    # 统计区间分布模式的条件概率
    patterns = []
    for rs in red_hist[:n_train]:
        z1c = sum(1 for r in rs if r <= 11)
        z2c = sum(1 for r in rs if 12 <= r <= 22)
        z3c = sum(1 for r in rs if r >= 23)
        patterns.append((z1c, z2c, z3c))
    if n_train < 10:
        return {n: 50 for n in ALL_RED}, {}
    # 最近10期最常见模式
    recent_p = Counter(patterns[-10:])
    top_pat = [p for p, _ in recent_p.most_common(3)]
    # 在这些模式下各号码的条件概率
    cond_hit = Counter()
    cond_tot = Counter()
    for i in range(n_train):
        if patterns[i] in top_pat:
            for r in red_hist[i]:
                cond_hit[r] += 1
            cond_tot["total"] += 1
    rs = {}
    for n in ALL_RED:
        rs[n] = cond_hit.get(n, 0) / max(cond_tot.get("total", 1), 1) * 100
    # 蓝球也用区间模式条件
    bs = {}
    return zonal_norm(rs), bs

# ── 6. 连号分析 ──
def m06_consecutive(red_hist, blue_hist, n_train):
    # 统计连号对 (a, a+1) 的出现频率
    pair_freq = Counter()
    for rs in red_hist[:n_train]:
        srs = sorted(rs)
        for i in range(len(srs)-1):
            if srs[i+1] - srs[i] == 1:
                pair_freq[(srs[i], srs[i+1])] += 1
    # 上期连号对的延续概率
    last = sorted(red_hist[n_train-1])
    last_pairs = set()
    for i in range(len(last)-1):
        if last[i+1] - last[i] == 1:
            last_pairs.add((last[i], last[i+1]))
    rs = Counter()
    for (a, b), cnt in pair_freq.items():
        rs[a] += cnt
        rs[b] += cnt
    # 上期连号的位置附近
    for (a, b) in last_pairs:
        for n in [a-1, a+1, b-1, b+1]:
            if 1 <= n <= 33:
                rs[n] += 30
    return zonal_norm(dict(rs)), {}

# ── 7. 跨度/间距分析 ──
def m07_gap(red_hist, blue_hist, n_train):
    # 分析相邻号码间距的分布
    gaps = Counter()
    for rs in red_hist[:n_train]:
        srs = sorted(rs)
        for i in range(len(srs)-1):
            gaps[srs[i+1] - srs[i]] += 1
    # 上期间距模式
    last = sorted(red_hist[n_train-1])
    last_gaps = []
    for i in range(len(last)-1):
        last_gaps.append(last[i+1] - last[i])
    # 寻找与上期间距模式最相似的历史期次
    rs = {n: 0 for n in ALL_RED}
    for i in range(n_train-1):
        srs = sorted(red_hist[i])
        hgaps = []
        for j in range(len(srs)-1):
            hgaps.append(srs[j+1]-srs[j])
        if hgaps == last_gaps:  # 完全相同的间距模式
            for r in red_hist[i+1]:
                rs[r] += 50
    return zonal_norm(rs), {}

# ── 8. 区间比模式 ──
def m08_zone_ratio(red_hist, blue_hist, n_train):
    zp = []
    for rs in red_hist[:n_train]:
        zp.append((
            sum(1 for r in rs if r <= 11),
            sum(1 for r in rs if 12 <= r <= 22),
            sum(1 for r in rs if r >= 23)
        ))
    if n_train < 5:
        return {n: 50 for n in ALL_RED}, {}
    # 最近5期的平均区间比
    recent = zp[-5:]
    avg_z1 = sum(p[0] for p in recent) / 5
    avg_z2 = sum(p[1] for p in recent) / 5
    avg_z3 = sum(p[2] for p in recent) / 5
    # 找到历史中最接近这个平均区间比的期次的下一期
    best_i = -1
    best_dist = float('inf')
    for i in range(n_train-1):
        p = zp[i]
        d = abs(p[0]-avg_z1) + abs(p[1]-avg_z2) + abs(p[2]-avg_z3)
        if d < best_dist:
            best_dist = d
            best_i = i
    rs = Counter()
    if best_i >= 0 and best_i + 1 < n_train:
        for r in red_hist[best_i+1]:
            rs[r] += 100
    return zonal_norm(dict(rs)), {}

# ── 9. 综合评分(原有9方法融合) ──
# 此方法本身是其他8方法的等权平均，作为对比基线
# 不做此方法，改为在main中手动融合对照

# ── 10.【V5】共现矩阵(跨期有向转移 + 种子自投排除) ──
def m10_cooccurrence(red_hist, blue_hist, n_train):
    """
    跨期共现矩阵：P(下期B | 上期A)
    - 短期窗口=12期（正交测试验证最优，共现关系会过时）
    - 等权(decay=1.0)（短窗口内无需衰减）
    - 只排除种子自投，不排除上期已出号码
    - 正交测试：42.92%(103/240)，距V5天花板仅差0.41%
    """
    if n_train < 3:
        return {n: 50 for n in ALL_RED}, {}
    
    WINDOW = 12  # 只看最近12期，共现关系短期有效
    
    trans = {a: Counter() for a in ALL_RED}
    total_w = {a: 0.0 for a in ALL_RED}
    
    start = max(1, n_train - WINDOW)
    for i in range(start, n_train):
        w = 1.0  # 短窗口等权
        pset = set(red_hist[i-1])
        cset = set(red_hist[i])
        for p in pset:
            for c in cset:
                if c != p:
                    trans[p][c] += w
                    total_w[p] += w
    
    last = red_hist[n_train-1]
    scores = Counter()
    
    for seed in last:
        tw = max(total_w.get(seed, 0), 1)
        for num, cnt in trans[seed].items():
            if num != seed:
                scores[num] += (cnt / tw) * 100
    
    for n in ALL_RED:
        if n not in scores:
            scores[n] = 0
    
    return zonal_norm(dict(scores)), {}

# ── 11.【V5】位序约束 ──
def m11_position_constraint(red_hist, blue_hist, n_train):
    """
    每个位置(第1-6个红球按升序)有独立分布范围
    位1: 01-10, 位2: 03-16, 位3: 08-22, 位4: 14-27, 位5: 19-31, 位6: 24-33
    方法：根据上期各位置号码，找历史上同位置号码的下一期出现区间
    """
    if n_train < 5:
        return {n: 50 for n in ALL_RED}, {}
    # 对每期排序
    sorted_hist = [sorted(rs) for rs in red_hist[:n_train]]
    # 统计每个位置上的号码分布
    pos_count = [Counter() for _ in range(6)]
    for srs in sorted_hist:
        for pos in range(6):
            pos_count[pos][srs[pos]] += 1
    # 基于各位置分布打分
    scores = Counter()
    for pos in range(6):
        for num, cnt in pos_count[pos].items():
            scores[num] += cnt * 100 / n_train
    return zonal_norm(dict(scores)), {}

# ── 12.【V5】指数衰减 ──
def m12_exponential_decay(red_hist, blue_hist, n_train, alpha=0.92):
    """
    指数衰减频率统计：越近的期次权重越大
    alpha=0.92 → 第1期权重0.92^49≈0.017, 最近期权重1.0
    比alpha=0.85更均衡，保留长期信号
    """
    if n_train == 0:
        return {n: 50 for n in ALL_RED}, {}
    rs = Counter()
    bs = Counter()
    for i in range(n_train):
        w = alpha ** (n_train - 1 - i)
        for r in red_hist[i]:
            rs[r] += w
        bs[blue_hist[i]] += w
    return zonal_norm(dict(rs)), norm01(dict(bs))

# ── 13.【V5】蓝球泊松间隔 ──
def m13_poisson_gap(red_hist, blue_hist, n_train):
    """
    蓝球间隔泊松建模 + 邻号叠加
    1. 统计每个蓝号的历史出现间隔
    2. 当前遗漏接近历史平均间隔 → 高分（号码"该出了"）
    3. 上期邻号大幅加权
    """
    if n_train < 3:
        return {n: 50 for n in ALL_RED}, {}
    
    last_pos = {n: -1 for n in ALL_BLUE}
    gaps = {n: [] for n in ALL_BLUE}
    for i in range(n_train):
        b = blue_hist[i]
        if last_pos[b] >= 0:
            gaps[b].append(i - last_pos[b])
        last_pos[b] = i
    
    bs = {}
    for n in ALL_BLUE:
        current_miss = n_train - 1 - last_pos.get(n, -1)
        if current_miss < 0:
            current_miss = n_train
        
        if gaps[n]:
            avg_gap = sum(gaps[n]) / len(gaps[n])
            std_gap = (sum((g - avg_gap) ** 2 for g in gaps[n]) / len(gaps[n])) ** 0.5
            # 距离期望间隔越近，得分越高
            z = abs(current_miss - avg_gap) / max(std_gap, 1)
            bs[n] = 100 / (1 + z)  # 标准分数越低越好
        else:
            bs[n] = 20  # 无历史数据，给低分
    
    # 上期邻号权重加倍
    lb = blue_hist[n_train-1]
    if lb > 1: bs[lb-1] = bs.get(lb-1, 0) * 4
    if lb < 16: bs[lb+1] = bs.get(lb+1, 0) * 4
    
    return {n: 50 for n in ALL_RED}, norm01(bs)

# ── 14.【新增】短期热度 ──
def m14_short_term_heat(red_hist, blue_hist, n_train, window=10):
    """
    短期热度：只看近N期的出现密度，完全不考虑长期频率
    - window：回顾窗口大小（5/10/15/20）
    - 红球：统计每个号码在最近window期内的出现次数，直接作为分数
    - 蓝球：同理
    """
    start = max(0, n_train - window)
    rcnt = Counter()
    for rs in red_hist[start:n_train]:
        rcnt.update(rs)
    bcnt = Counter()
    for b in blue_hist[start:n_train]:
        bcnt[b] += 1
    # 归一化
    rmax = max(rcnt.values()) if rcnt else 1
    rs = {n: rcnt.get(n, 0) / rmax * 100 for n in ALL_RED}
    bmax = max(bcnt.values()) if bcnt else 1
    bs = {n: bcnt.get(n, 0) / bmax * 100 for n in ALL_BLUE}
    return rs, bs
# ── 15.【新增】定期洗牌分区选方法 ──
# 模块级状态（因为每期独立调用，需要跨期保持状态）
_periodic_state = {
    "z1_method": "01_频率分析",
    "z2_method": "01_频率分析", 
    "z3_method": "01_频率分析",
}

def m15_periodic_reset_zone(red_hist, blue_hist, n_train, 
                             reset_interval=8, eval_window=10, start_test=10):
    """
    定期洗牌分区选方法（N=8,W=10，正交验证最优）
    每8期重新评估各方法在每区近10期的表现，选出每区最优方法
    非洗牌期沿用上轮选择
    """
    if n_train < start_test:
        return {n: 50 for n in ALL_RED}, {}
    
    periods_since_start = n_train - start_test
    need_reset = (periods_since_start % reset_interval == 0)
    
    if need_reset:
        # 评估窗口：最近 eval_window 期（已发生过的测试期）
        eval_from = max(start_test, n_train - eval_window)
        if eval_from >= n_train:
            eval_from = max(0, start_test - eval_window)
            eval_end = start_test - 1
        else:
            eval_end = n_train - 1
        
        # 统计各方法在评估窗口内每区命中数（排除自身避免递归）
        zone_hits = {}
        for name, func in RED_METHODS:
            if name == "15_定期洗牌分区":
                continue
            zh = {"Z1": 0, "Z2": 0, "Z3": 0}
            for past_idx in range(eval_from, eval_end + 1):
                if past_idx >= start_test:
                    _, past_reds, _ = RAW_DATA[past_idx]
                    pz1 = [r for r in past_reds if r <= 11]
                    pz2 = [r for r in past_reds if 12 <= r <= 22]
                    pz3 = [r for r in past_reds if r >= 23]
                    r_scores, _ = func(red_hist, blue_hist, past_idx)
                    zp1 = top_n_in_zone(r_scores, Z1, 4)
                    zp2 = top_n_in_zone(r_scores, Z2, 4)
                    zp3 = top_n_in_zone(r_scores, Z3, 4)
                    zh["Z1"] += len(set(zp1) & set(pz1))
                    zh["Z2"] += len(set(zp2) & set(pz2))
                    zh["Z3"] += len(set(zp3) & set(pz3))
            zone_hits[name] = zh
        
        _periodic_state["z1_method"] = max(zone_hits, key=lambda n: zone_hits[n]["Z1"])
        _periodic_state["z2_method"] = max(zone_hits, key=lambda n: zone_hits[n]["Z2"])
        _periodic_state["z3_method"] = max(zone_hits, key=lambda n: zone_hits[n]["Z3"])
    
    # 用选中的方法分别预测各区
    z1_func = next(f for n, f in RED_METHODS if n == _periodic_state["z1_method"])
    z2_func = next(f for n, f in RED_METHODS if n == _periodic_state["z2_method"])
    z3_func = next(f for n, f in RED_METHODS if n == _periodic_state["z3_method"])
    
    z1_scores, _ = z1_func(red_hist, blue_hist, n_train)
    z2_scores, _ = z2_func(red_hist, blue_hist, n_train)
    z3_scores, _ = z3_func(red_hist, blue_hist, n_train)
    
    # 合并：每区只用选中方法的分数
    merged = {}
    for n in Z1: merged[n] = z1_scores.get(n, 0)
    for n in Z2: merged[n] = z2_scores.get(n, 0)
    for n in Z3: merged[n] = z3_scores.get(n, 0)
    
    return zonal_norm(merged), {}

# ============================================================
# 所有方法注册（红球方法 + 蓝球方法）
# ============================================================
RED_METHODS = [
    ("01_频率分析",      lambda rh, bh, nt: m01_frequency(rh, bh, nt)),
    ("02_遗漏分析",      lambda rh, bh, nt: m02_missing(rh, bh, nt)),
    ("03_邻号分析",      lambda rh, bh, nt: m03_adjacent(rh, bh, nt)),
    ("04_马尔可夫",      lambda rh, bh, nt: m04_markov(rh, bh, nt)),
    ("05_贝叶斯",       lambda rh, bh, nt: m05_bayesian(rh, bh, nt)),
    ("06_连号分析",      lambda rh, bh, nt: m06_consecutive(rh, bh, nt)),
    ("07_间距分析",      lambda rh, bh, nt: m07_gap(rh, bh, nt)),
    ("08_区间比模式",    lambda rh, bh, nt: m08_zone_ratio(rh, bh, nt)),
    ("10_共现矩阵",      lambda rh, bh, nt: m10_cooccurrence(rh, bh, nt)),
    ("11_位序约束",      lambda rh, bh, nt: m11_position_constraint(rh, bh, nt)),
    ("12_指数衰减",      lambda rh, bh, nt: m12_exponential_decay(rh, bh, nt)),
    ("14_短期热度20",    lambda rh, bh, nt: m14_short_term_heat(rh, bh, nt, window=20)),
    ("15_定期洗牌分区",  lambda rh, bh, nt: m15_periodic_reset_zone(rh, bh, nt)),
]

BLUE_METHODS = [
    ("03_蓝球邻号",      lambda rh, bh, nt: (m03_adjacent(rh, bh, nt)[0], m03_adjacent(rh, bh, nt)[1])),
    ("13_泊松间隔蓝球",  lambda rh, bh, nt: (m13_poisson_gap(rh, bh, nt)[0], m13_poisson_gap(rh, bh, nt)[1])),
    ("14_短期热度蓝球",  lambda rh, bh, nt: (m14_short_term_heat(rh, bh, nt, window=20)[0], m14_short_term_heat(rh, bh, nt, window=20)[1])),
]

# ============================================================
# 回测引擎
# ============================================================
def run_backtest():
    all_reds = [reds for _, reds, _ in RAW_DATA]
    all_blues = [blue for _, _, blue in RAW_DATA]

    red_results = {name: [] for name, _ in RED_METHODS}
    blue_results = {name: [] for name, _ in BLUE_METHODS}

    # 42期回测 (test_idx from 10 to 51 = period 2026027 ~ 2026068)
    for test_idx in range(10, N):
        period, actual_reds, actual_blue = RAW_DATA[test_idx]
        actual = set(actual_reds)
        z1a = [r for r in actual_reds if r <= 11]
        z2a = [r for r in actual_reds if 12 <= r <= 22]
        z3a = [r for r in actual_reds if r >= 23]

        for name, func in RED_METHODS:
            r_scores, b_scores = func(all_reds, all_blues, test_idx)
            z1p = top_n_in_zone(r_scores, Z1, 4)
            z2p = top_n_in_zone(r_scores, Z2, 4)
            z3p = top_n_in_zone(r_scores, Z3, 4)
            h1 = len(set(z1p) & set(z1a))
            h2 = len(set(z2p) & set(z2a))
            h3 = len(set(z3p) & set(z3a))
            red_results[name].append({
                "period": period, "hits": h1+h2+h3,
                "z1_hit": h1, "z2_hit": h2, "z3_hit": h3,
                "z1_pred": z1p, "z2_pred": z2p, "z3_pred": z3p,
                "z1_act": z1a, "z2_act": z2a, "z3_act": z3a,
            })

        for name, func in BLUE_METHODS:
            r_scores, b_scores = func(all_reds, all_blues, test_idx)
            bp = top_n_blue(b_scores, 4)
            bh = 1 if actual_blue in bp else 0
            blue_results[name].append({
                "period": period, "hit": bh,
                "pred": bp, "actual": actual_blue,
            })

    return red_results, blue_results

# ============================================================
# 统计 & 输出
# ============================================================
def compute_red_stats(results):
    n = len(results)
    total_hits = sum(r["hits"] for r in results)
    z1_hits = sum(r["z1_hit"] for r in results)
    z2_hits = sum(r["z2_hit"] for r in results)
    z3_hits = sum(r["z3_hit"] for r in results)
    return {
        "count": n,
        "total_rate": total_hits / (n * 6) * 100,   # 分母=6（每期实际出6个红球）
        "z1_rate": z1_hits / (n * 4) * 100,
        "z2_rate": z2_hits / (n * 4) * 100,
        "z3_rate": z3_hits / (n * 4) * 100,
        "avg_per_period": total_hits / n,
    }

def compute_blue_stats(results):
    n = len(results)
    hits = sum(r["hit"] for r in results)
    return {"count": n, "rate": hits / n * 100}

def main():
    print("=" * 75)
    print("  S1 双色球预测系统 — 14方法 42期独立回测 (含定期洗牌)")
    print("=" * 75)

    red_results, blue_results = run_backtest()

    # ── 红球排名 ──
    print(f"\n  {'─' * 70}")
    print(f"  {'红球方法 (4选/区)':<22} {'总命中':>6} {'一区':>6} {'二区':>6} {'三区':>6} {'均中':>5}")
    print(f"  {'─' * 70}")

    red_ranking = []
    for name, rows in red_results.items():
        s = compute_red_stats(rows)
        red_ranking.append((name, s))
    red_ranking.sort(key=lambda x: -x[1]["total_rate"])

    for name, s in red_ranking:
        bar = "#" * int(s["total_rate"] / 2)
        print(f"  {name:<22} {s['total_rate']:>5.1f}% {s['z1_rate']:>5.1f}% {s['z2_rate']:>5.1f}% {s['z3_rate']:>5.1f}% {s['avg_per_period']:>4.2f}  {bar}")

    # ── 蓝球排名 ──
    print(f"\n  {'─' * 70}")
    print(f"  {'蓝球方法 (4选)':<22} {'命中率':>8}")
    print(f"  {'─' * 70}")

    blue_ranking = []
    for name, rows in blue_results.items():
        s = compute_blue_stats(rows)
        blue_ranking.append((name, s))
    blue_ranking.sort(key=lambda x: -x[1]["rate"])

    for name, s in blue_ranking:
        bar = "#" * int(s["rate"] / 2)
        print(f"  {name:<22} {s['rate']:>7.1f}%  {bar}")

    # ── 第9方法：其他8方法等权融合对照 ──
    print(f"\n  {'─' * 70}")
    print(f"  {'对照':<22} {'红球':>6} {'一区':>6} {'二区':>6} {'三区':>6} {'蓝球':>7} {'均中':>5}")

    # V5基准
    benchmarks = [
        ("V5共现独立(基准)",  43.3, 42.5, 44.3, 43.3, 27.5, 2.60),
        ("V5邻号蓝球(基准)",  39.6, 36.2, 40.0, 42.2, 40.0, 2.38),
        ("V6融合(基准)",      41.2, 40.0, 40.0, 43.3, 35.0, 2.48),
    ]
    for name, rr, z1r, z2r, z3r, br, ap in benchmarks:
        print(f"  {name:<22} {rr:>5.1f}% {z1r:>5.1f}% {z2r:>5.1f}% {z3r:>5.1f}% {br:>6.1f}% {ap:>4.2f}")

    # S1最佳红球
    if red_ranking:
        best_r = red_ranking[0]
        print(f"  S1最佳红: {best_r[0]:<14} {best_r[1]['total_rate']:>5.1f}% {best_r[1]['z1_rate']:>5.1f}% {best_r[1]['z2_rate']:>5.1f}% {best_r[1]['z3_rate']:>5.1f}%        {best_r[1]['avg_per_period']:>4.2f}")
    if blue_ranking:
        best_b = blue_ranking[0]
        print(f"  S1最佳蓝: {best_b[0]:<14}                                  {best_b[1]['rate']:>6.1f}%")

    # ── Excel导出 ──
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        top_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def write_header(ws, row, headers):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = Alignment(horizontal='center')
                c.border = thin_border

        # Sheet1: 排名汇总
        ws1 = wb.active
        ws1.title = "方法排名"
        ws1.cell(row=1, column=1, value="S1 双色球预测系统 — 14方法独立回测").font = Font(bold=True, size=14)
        ws1.cell(row=2, column=1, value="回测范围: 2026027~2026068 (42期) | 每区4选").font = Font(italic=True)

        write_header(ws1, 4, ["排名", "红球方法", "总命中%", "一区%", "二区%", "三区%", "每期均中"])
        for ri, (name, s) in enumerate(red_ranking, 5):
            for ci, v in enumerate([ri-4, name, round(s["total_rate"],1), round(s["z1_rate"],1),
                                     round(s["z2_rate"],1), round(s["z3_rate"],1),
                                     round(s["avg_per_period"],2)], 1):
                c = ws1.cell(row=ri, column=ci, value=v)
                c.border = thin_border
                c.alignment = Alignment(horizontal='center')
                if ri == 5:
                    c.fill = top_fill
                    c.font = Font(bold=True)

        row_b = 5 + len(red_ranking) + 1
        write_header(ws1, row_b, ["排名", "蓝球方法", "命中率%", "", "", "", ""])
        for ri, (name, s) in enumerate(blue_ranking, row_b+1):
            for ci, v in enumerate([ri-row_b, name, round(s["rate"],1), "", "", "", ""], 1):
                c = ws1.cell(row=ri, column=ci, value=v)
                c.border = thin_border
                c.alignment = Alignment(horizontal='center')
                if ri == row_b+1:
                    c.fill = top_fill
                    c.font = Font(bold=True)

        # Sheet2: 最佳方法逐期明细
        if red_ranking:
            ws2 = wb.create_sheet(f"{red_ranking[0][0]}_明细")
            write_header(ws2, 1, ["期号","一区推","一区实","中","二区推","二区实","中","三区推","三区实","中","总中"])
            for ri, r in enumerate(red_results[red_ranking[0][0]], 2):
                ws2.cell(row=ri, column=1, value=r["period"]).border = thin_border
                ws2.cell(row=ri, column=2, value=" ".join(f"{n:02d}" for n in r["z1_pred"])).border = thin_border
                ws2.cell(row=ri, column=3, value=" ".join(f"{n:02d}" for n in r["z1_act"])).border = thin_border
                ws2.cell(row=ri, column=4, value=r["z1_hit"]).border = thin_border
                ws2.cell(row=ri, column=5, value=" ".join(f"{n:02d}" for n in r["z2_pred"])).border = thin_border
                ws2.cell(row=ri, column=6, value=" ".join(f"{n:02d}" for n in r["z2_act"])).border = thin_border
                ws2.cell(row=ri, column=7, value=r["z2_hit"]).border = thin_border
                ws2.cell(row=ri, column=8, value=" ".join(f"{n:02d}" for n in r["z3_pred"])).border = thin_border
                ws2.cell(row=ri, column=9, value=" ".join(f"{n:02d}" for n in r["z3_act"])).border = thin_border
                ws2.cell(row=ri, column=10, value=r["z3_hit"]).border = thin_border
                ws2.cell(row=ri, column=11, value=r["hits"]).border = thin_border

        # Sheet3: 当前预测 (2026069)
        ws3 = wb.create_sheet("S1当前预测")
        all_r = [reds for _, reds, _ in RAW_DATA]
        all_b = [blue for _, _, blue in RAW_DATA]
        ws3.cell(row=1, column=1, value="S1系统 — 2026069期预测").font = Font(bold=True, size=14)
        ws3.cell(row=2, column=1, value=f"上期(2026068): 红[03,05,16,18,29,32] 蓝04").font = Font(italic=True)

        row = 4
        write_header(ws3, row, ["方法", "一区推4", "二区推4", "三区推4", "蓝推4"])
        for name, func in RED_METHODS:
            row += 1
            rs, bs = func(all_r, all_b, N)
            z1p = top_n_in_zone(rs, Z1, 4)
            z2p = top_n_in_zone(rs, Z2, 4)
            z3p = top_n_in_zone(rs, Z3, 4)
            bp = top_n_blue(bs, 4)
            blue_str = "无蓝球分析" if not bp else " ".join(f"{n:02d}" for n in bp)
            for ci, v in enumerate([name, " ".join(f"{n:02d}" for n in z1p),
                                     " ".join(f"{n:02d}" for n in z2p),
                                     " ".join(f"{n:02d}" for n in z3p),
                                     blue_str], 1):
                ws3.cell(row=row, column=ci, value=v).border = thin_border

        fpath = "C:\\Users\\27218\\WorkBuddy\\2026-06-12-16-55-40\\双色球S1回测报告.xlsx"
        wb.save(fpath)
        print(f"\n  Excel已保存: 双色球S1回测报告.xlsx")
    except Exception as e:
        print(f"\n  Excel保存失败: {e}")

if __name__ == "__main__":
    main()
